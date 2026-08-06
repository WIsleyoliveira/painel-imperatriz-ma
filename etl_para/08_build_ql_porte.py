"""QL_Grupo_Empresas_por_Porte_{Cidade}.xlsx — réplica do conceito do
QL_Grupo_Empresas_por_Porte.xlsx de Imperatriz, adaptado para o esquema de
porte disponível aqui (MEI vs Demais, sem ME/EPP x Grande/Médio separados).
Sheets: Totais por Porte e Ano | {CIDADE}_MEI | {CIDADE}_Demais | QL Geral - vs PA"""

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

STAGING = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/staging"
REFS = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/refs"
OUT_DIR = "/Volumes/WISKET/painel-imperatriz-ma-main"

ANOS = [2001, 2003, 2005, 2007, 2009, 2011, 2013, 2015, 2017, 2019, 2021, 2023, 2025]
MUNIC_CODES = {"Ananindeua": "0415", "Capanema": "0443"}
REF = "Pará"
PARA_BUCKETS = ["Ananindeua", "Capanema", "ParaResto"]

mapa = pd.read_csv(f"{REFS}/_mapa_grupo_secao_oficial.csv")
mapa["Grupo"] = mapa["Grupo"].astype(int)
grupo_lookup = mapa.set_index("Grupo")[["Seção", "Nome Seção", "Divisão", "Nome Divisão", "Nome Grupo"]].to_dict("index")

con = duckdb.connect()
con.execute("PRAGMA threads=6")

year_cols_ativas = ",\n".join(
    f"SUM(CASE WHEN ano_abertura<={y} AND is_ativa THEN 1 ELSE 0 END) AS ativas_{y}" for y in ANOS
)

print("Agregando por grupo x porte...", flush=True)
by_grupo_porte = con.execute(f"""
    SELECT
        CASE WHEN municipio='0415' AND uf='PA' THEN 'Ananindeua'
             WHEN municipio='0443' AND uf='PA' THEN 'Capanema'
             WHEN uf='PA' THEN 'ParaResto'
             ELSE 'Outro' END AS geo_bucket,
        grupo_codigo,
        CASE WHEN is_mei THEN 'MEI' ELSE 'Demais' END AS porte,
        {year_cols_ativas}
    FROM read_parquet('{STAGING}/estabele_staging_v2_matriz.parquet')
    WHERE grupo_codigo IS NOT NULL AND uf='PA'
    GROUP BY 1, 2, 3
""").df()
by_grupo_porte["grupo_int"] = by_grupo_porte["grupo_codigo"].astype(int)
val_cols = [f"ativas_{y}" for y in ANOS]

bold = Font(bold=True)
title_font = Font(bold=True, size=12)


def geo_sum(df_, buckets, group_cols):
    sub = df_[df_["geo_bucket"].isin(buckets)]
    return sub.groupby(group_cols, as_index=False)[val_cols].sum()


def calc_metrics(vals, total_last_year):
    representat = (vals[-1] / total_last_year * 100) if total_last_year else None
    cresc5 = (vals[-1] - vals[-4]) / vals[-4] * 100 if vals[-4] else None
    transicoes = sum(1 for i in range(1, len(vals)) if vals[i] > vals[i - 1])
    return representat, cresc5, f"{transicoes}/{len(vals)-1}"


def sheet_totais(wb, cidade):
    ws = wb.create_sheet("Totais por Porte e Ano")
    ws.cell(1, 1, f"TOTAL DE EMPRESAS POR PORTE — {cidade.upper()} E {REF.upper()} — 2001 A 2025").font = title_font
    ws.cell(2, 1, "Matriz ativa hoje, por data de abertura (corte 31/12 de cada ano). Fonte: CNPJ Receita Federal.")
    ws.cell(3, 1, "Nível").font = bold
    ws.cell(3, 2, "Porte").font = bold
    for j, a in enumerate(ANOS):
        ws.cell(3, 3 + j, a).font = bold

    cidade_pp = by_grupo_porte[by_grupo_porte.geo_bucket == cidade].groupby("porte")[val_cols].sum()
    para_pp = geo_sum(by_grupo_porte, PARA_BUCKETS, ["porte"]).set_index("porte")

    r = 4
    for nivel, dfx in [(cidade, cidade_pp), (REF, para_pp)]:
        primeiro = True
        totals = [0] * len(ANOS)
        for porte in ["MEI", "Demais"]:
            vals = [int(dfx.loc[porte, f"ativas_{a}"]) if porte in dfx.index else 0 for a in ANOS]
            ws.cell(r, 1, nivel if primeiro else "")
            ws.cell(r, 2, porte)
            for j, v in enumerate(vals):
                ws.cell(r, 3 + j, v)
                totals[j] += v
            primeiro = False
            r += 1
        ws.cell(r, 1, "")
        ws.cell(r, 2, "Total").font = bold
        for j, t in enumerate(totals):
            ws.cell(r, 3 + j, t).font = bold
        r += 1

    for j in range(1, 3 + len(ANOS)):
        ws.column_dimensions[get_column_letter(j)].width = 12


def sheet_porte_grupo(wb, cidade, porte):
    ws = wb.create_sheet(f"{cidade}_{porte}")
    ws.cell(1, 1, f"EMPRESAS {porte.upper()} POR SEÇÃO/DIVISÃO/GRUPO CNAE — {cidade.upper()} — 2001 A 2025").font = title_font

    header = ["Seção", "Nome Seção", "Divisão", "Nome Divisão", "Grupo", "Nome Grupo"] + \
        [a for a in ANOS] + [f"QL {porte}\n({cidade[:3]} vs {REF[:2]})", "QL Geral\n(todos portes)",
        "Representat. (%)", "Cresc. 5 anos (%)", "Consistência (0-12)"]
    for j, h in enumerate(header):
        ws.cell(3, j + 1, h).font = bold

    cidade_df = by_grupo_porte[(by_grupo_porte.geo_bucket == cidade) & (by_grupo_porte.porte == porte)]
    cidade_piv = cidade_df.groupby("grupo_int")[val_cols].sum()
    para_df = geo_sum(by_grupo_porte[by_grupo_porte.porte == porte], PARA_BUCKETS, ["grupo_int"])
    para_piv = para_df.set_index("grupo_int")

    cidade_all_portes = by_grupo_porte[by_grupo_porte.geo_bucket == cidade].groupby("grupo_int")[val_cols].sum()
    para_all_portes = geo_sum(by_grupo_porte, PARA_BUCKETS, ["grupo_int"]).set_index("grupo_int")

    cidade_total_by_year = cidade_piv.sum(axis=0)
    para_total_by_year = para_piv.sum(axis=0) if len(para_piv) else pd.Series(0, index=val_cols)
    cidade_geral_total_by_year = cidade_all_portes.sum(axis=0)
    para_geral_total_by_year = para_all_portes.sum(axis=0)

    r = 4
    for idx in sorted(cidade_piv.index):
        info = grupo_lookup.get(idx, {})
        vals = [int(cidade_piv.loc[idx, f"ativas_{a}"]) for a in ANOS]
        if sum(vals) == 0:
            continue
        ref_vals = [int(para_piv.loc[idx, f"ativas_{a}"]) if idx in para_piv.index else 0 for a in ANOS]
        ql_last = None
        city_tot = cidade_total_by_year[f"ativas_{ANOS[-1]}"]
        ref_tot = para_total_by_year[f"ativas_{ANOS[-1]}"]
        if city_tot and ref_vals[-1] and ref_tot:
            ql_last = (vals[-1] / city_tot) / (ref_vals[-1] / ref_tot)

        ql_geral = None
        gv = int(cidade_all_portes.loc[idx, f"ativas_{ANOS[-1]}"]) if idx in cidade_all_portes.index else 0
        grv = int(para_all_portes.loc[idx, f"ativas_{ANOS[-1]}"]) if idx in para_all_portes.index else 0
        cgt = cidade_geral_total_by_year[f"ativas_{ANOS[-1]}"]
        rgt = para_geral_total_by_year[f"ativas_{ANOS[-1]}"]
        if cgt and grv and rgt:
            ql_geral = (gv / cgt) / (grv / rgt)

        representat, cresc5, consist = calc_metrics(vals, city_tot)

        c = 1
        ws.cell(r, c, info.get("Seção", "")); c += 1
        ws.cell(r, c, info.get("Nome Seção", "")); c += 1
        ws.cell(r, c, info.get("Divisão", "")); c += 1
        ws.cell(r, c, info.get("Nome Divisão", "")); c += 1
        ws.cell(r, c, idx); c += 1
        ws.cell(r, c, info.get("Nome Grupo", "")); c += 1
        for v in vals:
            ws.cell(r, c, v); c += 1
        ws.cell(r, c, round(ql_last, 2) if ql_last is not None else "—"); c += 1
        ws.cell(r, c, round(ql_geral, 2) if ql_geral is not None else "—"); c += 1
        ws.cell(r, c, round(representat, 3) if representat is not None else "—"); c += 1
        ws.cell(r, c, round(cresc5, 1) if cresc5 is not None else "—"); c += 1
        ws.cell(r, c, consist); c += 1
        r += 1

    for j in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13


def sheet_ql_geral(wb, cidade):
    ws = wb.create_sheet(f"QL Geral - vs {REF[:2]}")
    ws.cell(1, 1, f"QL GERAL (TODOS OS PORTES) POR SEÇÃO/DIVISÃO/GRUPO — {cidade.upper()} vs {REF.upper()}").font = title_font
    header = ["Seção", "Nome Seção", "Divisão", "Nome Divisão", "Grupo", "Nome Grupo"] + \
        [f"QL {a}" for a in ANOS] + ["QL Médio", "Períodos Cresc."]
    for j, h in enumerate(header):
        ws.cell(2, j + 1, h).font = bold

    cidade_all = by_grupo_porte[by_grupo_porte.geo_bucket == cidade].groupby("grupo_int")[val_cols].sum()
    para_all = geo_sum(by_grupo_porte, PARA_BUCKETS, ["grupo_int"]).set_index("grupo_int")
    cidade_tot_yr = cidade_all.sum(axis=0)
    para_tot_yr = para_all.sum(axis=0)

    r = 3
    for idx in sorted(cidade_all.index):
        info = grupo_lookup.get(idx, {})
        vals = [int(cidade_all.loc[idx, f"ativas_{a}"]) for a in ANOS]
        if sum(vals) == 0:
            continue
        ref_vals = [int(para_all.loc[idx, f"ativas_{a}"]) if idx in para_all.index else 0 for a in ANOS]
        qls = []
        for v, rv, a in zip(vals, ref_vals, ANOS):
            ct, rt = cidade_tot_yr[f"ativas_{a}"], para_tot_yr[f"ativas_{a}"]
            qls.append((v / ct) / (rv / rt) if ct and rv and rt else None)
        validos = [q for q in qls if q is not None]
        ql_medio = sum(validos) / len(validos) if validos else None
        periodos = sum(1 for i in range(1, len(qls)) if qls[i] is not None and qls[i - 1] is not None and qls[i] > qls[i - 1])

        c = 1
        ws.cell(r, c, info.get("Seção", "")); c += 1
        ws.cell(r, c, info.get("Nome Seção", "")); c += 1
        ws.cell(r, c, info.get("Divisão", "")); c += 1
        ws.cell(r, c, info.get("Nome Divisão", "")); c += 1
        ws.cell(r, c, idx); c += 1
        ws.cell(r, c, info.get("Nome Grupo", "")); c += 1
        for q in qls:
            ws.cell(r, c, round(q, 2) if q is not None else "—"); c += 1
        ws.cell(r, c, round(ql_medio, 2) if ql_medio is not None else "—"); c += 1
        ws.cell(r, c, periodos); c += 1
        r += 1

    for j in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13


for cidade in MUNIC_CODES:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_totais(wb, cidade)
    sheet_porte_grupo(wb, cidade, "MEI")
    sheet_porte_grupo(wb, cidade, "Demais")
    sheet_ql_geral(wb, cidade)
    path = f"{OUT_DIR}/QL_Grupo_Empresas_por_Porte_{cidade}.xlsx"
    wb.save(path)
    print("Gravado:", path)

print("CONCLUIDO")
