"""Monta Estoque_Empregos_RAIS_Ananindeua_Capanema.xlsx no mesmo estilo do
arquivo usado para Imperatriz: uma aba por geografia (Ananindeua, Capanema,
Pará), Seção CNAE x Descrição x Empregos por ano x Δ% acumulado, com linha
TOTAL no final. Pará = soma de Ananindeua + Capanema + demais municípios."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

STAGING = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/staging"
REFS = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/refs"
OUT_XLSX = "/Volumes/WISKET/painel-imperatriz-ma-main/Estoque_Empregos_RAIS_Ananindeua_Capanema.xlsx"

ANOS = [2001, 2003, 2005, 2007, 2009, 2011, 2013, 2015, 2017, 2019, 2021, 2023, 2025]

df = pd.read_parquet(f"{STAGING}/rais_vinculos_agregado.parquet")
secoes = pd.read_csv(f"{REFS}/cnae_secoes.csv", dtype=str)
secoes = secoes.set_index("codigo")["nome"].to_dict()

df = df.dropna(subset=["secao_codigo"])

# Pará = soma das 3 geografias (estado inteiro)
para_total = df.groupby(["ano", "secao_codigo"], as_index=False)["empregos"].sum()
para_total["geo"] = "Pará"

df["geo"] = df["municipio"].map({"150080": "Ananindeua", "150293": "Capanema"}).fillna("__resto__")
cidades = df[df["geo"] != "__resto__"].groupby(["geo", "ano", "secao_codigo"], as_index=False)["empregos"].sum()

full = pd.concat([cidades, para_total[["geo", "ano", "secao_codigo", "empregos"]]], ignore_index=True)

wb = Workbook()
wb.remove(wb.active)

bold = Font(bold=True)
title_font = Font(bold=True, size=12)

for geo in ["Ananindeua", "Capanema", "Pará"]:
    ws = wb.create_sheet(geo)
    sub = full[full["geo"] == geo]
    pivot = sub.pivot_table(index="secao_codigo", columns="ano", values="empregos", aggfunc="sum").fillna(0)
    pivot = pivot.reindex(columns=ANOS, fill_value=0)

    ws.cell(1, 1, f"ESTOQUE DE EMPREGOS FORMAIS POR SEÇÃO CNAE — {geo.upper()} — 2001 A 2025").font = title_font
    ws.cell(2, 1, "Vínculos empregatícios ativos em 31/12. Fonte: RAIS/MTE — microdados de vínculos "
                  "(2001-2005: CNAE 95 convertido via tabela de correspondência oficial IBGE; 2007+: CNAE 2.0 direto).")

    header_row = 3
    ws.cell(header_row, 1, "Seção").font = bold
    ws.cell(header_row, 2, "Descrição").font = bold
    for j, ano in enumerate(ANOS):
        ws.cell(header_row, 3 + j, f"Empregos\n{ano}").font = bold
    ws.cell(header_row, 3 + len(ANOS), "Δ% acum.\n2001→2025").font = bold

    r = header_row + 1
    totals = [0] * len(ANOS)
    for secao in sorted(pivot.index):
        vals = [int(pivot.loc[secao, ano]) for ano in ANOS]
        ws.cell(r, 1, secao)
        ws.cell(r, 2, secoes.get(secao, secao))
        for j, v in enumerate(vals):
            ws.cell(r, 3 + j, v)
            totals[j] += v
        primeiro = next((v for v in vals if v > 0), None)
        ultimo = vals[-1]
        if primeiro:
            delta = (ultimo - primeiro) / primeiro * 100
            ws.cell(r, 3 + len(ANOS), f"{delta:.1f}%")
        else:
            ws.cell(r, 3 + len(ANOS), "—")
        r += 1

    ws.cell(r, 1, "TOTAL").font = bold
    for j, t in enumerate(totals):
        ws.cell(r, 3 + j, t).font = bold
    delta_total = (totals[-1] - totals[0]) / totals[0] * 100 if totals[0] else None
    ws.cell(r, 3 + len(ANOS), f"{delta_total:.1f}%" if delta_total is not None else "—").font = bold

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    for j in range(len(ANOS) + 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(3 + j)].width = 13

wb.save(OUT_XLSX)
print("Gravado:", OUT_XLSX)
for geo in ["Ananindeua", "Capanema", "Pará"]:
    sub = full[full["geo"] == geo]
    print(geo, "total 2025:", int(sub[sub.ano == 2025]["empregos"].sum()))
