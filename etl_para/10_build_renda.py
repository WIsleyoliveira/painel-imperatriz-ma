"""Renda_por_Secao/Divisao/Grupo_CNAE_RAIS_{Cidade}.xlsx — remuneração média
nominal (Vl Rem Média Nom) dos vínculos ativos em 2025, Ananindeua/Capanema
vs Pará. QL Renda = renda cidade / renda Pará."""

import csv
import io
import subprocess
import time
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DOWNLOADS = Path("/Users/wisley/Downloads")
WORK_DIR = Path("/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/staging")
REFS_DIR = Path("/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/refs")
OUT_DIR = "/Volumes/WISKET/painel-imperatriz-ma-main"

MUNIC_CODES = {"Ananindeua": "150080", "Capanema": "150293"}
REF = "Pará"
ARCHIVE = DOWNLOADS / "RAIS_VINC_PUB_NORTE2025.7z"

mapa = pd.read_csv(f"{REFS_DIR}/_mapa_grupo_secao_oficial.csv")
mapa["Grupo"] = mapa["Grupo"].astype(int)
mapa["Divisão"] = mapa["Divisão"].astype(int)
grupo_lookup = mapa.set_index("Grupo")[["Seção", "Nome Seção", "Divisão", "Nome Divisão", "Nome Grupo"]].to_dict("index")
secoes_unicas = mapa[["Seção", "Nome Seção"]].drop_duplicates().set_index("Seção")["Nome Seção"].to_dict()
divisoes_unicas = mapa[["Seção", "Nome Seção", "Divisão", "Nome Divisão"]].drop_duplicates()


def get_header_and_delim(archive_path):
    proc = subprocess.run(["7z", "e", "-y", str(archive_path), "-so"],
                           stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, check=True)
    first_line = proc.stdout[:40000].split(b"\n")[0]
    text = first_line.decode("latin-1")
    delim = "," if text.count(",") > text.count(";") else ";"
    header = next(csv.reader(io.StringIO(text), delimiter=delim))
    return header, delim


def find_col(header, *needles, exclude=()):
    for i, h in enumerate(header):
        h_norm = h.strip().lower()
        if all(n.lower() in h_norm for n in needles) and not any(e.lower() in h_norm for e in exclude):
            return i
    raise ValueError(f"Coluna não encontrada para {needles} em {header}")


def main():
    t0 = time.time()
    header, delim = get_header_and_delim(ARCHIVE)
    idx_ativo = find_col(header, "vínculo ativo")
    idx_munic = find_col(header, "município", exclude=("trab",))
    idx_cnae = find_col(header, "cnae 2.0", "subclasse")
    idx_renda = find_col(header, "vl rem média nom")
    print(f"Colunas: ativo={idx_ativo} municipio={idx_munic} cnae={idx_cnae} renda={idx_renda} delim={delim!r}", flush=True)

    txt_path = WORK_DIR / "tmp_renda_2025.txt"
    print("Extraindo (com iconv, mesmo bug de encoding do CNPJ pode ocorrer)...", flush=True)
    with open(txt_path, "wb") as f:
        p7z = subprocess.Popen(["7z", "e", "-y", str(ARCHIVE), "-so"], stdout=subprocess.PIPE, stderr=subprocess.DEVNULL)
        subprocess.run(["iconv", "-f", "ISO-8859-1", "-t", "UTF-8"], stdin=p7z.stdout, stdout=f, check=True)
        p7z.wait()

    con = duckdb.connect()
    con.execute("PRAGMA threads=6")
    con.execute(f"""
        CREATE VIEW mapa AS
        SELECT CAST("Grupo" AS VARCHAR) AS grupo_codigo, "Seção" AS secao_codigo
        FROM read_csv('{REFS_DIR}/_mapa_grupo_secao_oficial.csv', header=True)
    """)

    col_names = [f"c{i}" for i in range(len(header))]
    cols_sql = ", ".join(f"'{c}':'VARCHAR'" for c in col_names)

    query = f"""
        SELECT
            CASE WHEN TRIM(c{idx_munic}) IN ('150080','150293') THEN TRIM(c{idx_munic}) ELSE 'PARA' END AS municipio,
            TRY_CAST(REPLACE(TRIM(c{idx_cnae}), ' ', '') AS INTEGER) AS cnae_raw,
            TRY_CAST(TRIM(c{idx_renda}) AS DOUBLE) AS renda
        FROM read_csv('{txt_path}', delim='{delim}', header=True, quote='"',
            columns={{{cols_sql}}}, strict_mode=False, ignore_errors=True)
        WHERE TRIM(c{idx_ativo}) = '1'
    """
    df = con.execute(query).df()
    txt_path.unlink()
    print(f"Extraído: {len(df):,} vínculos ativos — {time.time()-t0:.0f}s", flush=True)

    df["grupo_int"] = (df["cnae_raw"] // 10000).astype("Int64")
    df = df.dropna(subset=["grupo_int", "renda"])
    df = df[df["renda"] > 0]
    df["grupo_int"] = df["grupo_int"].astype(int)

    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)

    def agg_by(keys):
        g = df.groupby(["municipio"] + keys).agg(renda_media=("renda", "mean"), vinculos=("renda", "count")).reset_index()
        return g

    def build_level(cidade, level):
        if level == "grupo":
            keys = ["grupo_int"]
        elif level == "divisao":
            df["divisao_int"] = df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Divisão"))
            keys = ["divisao_int"]
        else:
            df["Seção"] = df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Seção"))
            keys = ["Seção"]
        g = agg_by(keys)
        cidade_g = g[g.municipio == MUNIC_CODES[cidade]].set_index(keys)
        ref_g = g[g.municipio.isin(["150080", "150293", "PARA"]) & (g.municipio != MUNIC_CODES[cidade])]
        # Pará = soma ponderada de TODOS os municípios (incl. a própria cidade)
        ref_all = df.copy()
        if level == "grupo":
            ref_keys = ["grupo_int"]
        elif level == "divisao":
            ref_keys = ["divisao_int"]
        else:
            ref_keys = ["Seção"]
        ref_agg = ref_all.groupby(ref_keys).agg(renda_media=("renda", "mean"), vinculos=("renda", "count"))

        rows = []
        idx_set = sorted(set(cidade_g.index) | set(ref_agg.index), key=str)
        for idx in idx_set:
            renda_c = cidade_g.loc[idx, "renda_media"] if idx in cidade_g.index else None
            vinc_c = int(cidade_g.loc[idx, "vinculos"]) if idx in cidade_g.index else 0
            renda_r = ref_agg.loc[idx, "renda_media"] if idx in ref_agg.index else None
            vinc_r = int(ref_agg.loc[idx, "vinculos"]) if idx in ref_agg.index else 0
            ql = (renda_c / renda_r) if (renda_c is not None and renda_r) else None
            rows.append({"idx": idx, "renda_c": renda_c, "vinc_c": vinc_c, "renda_r": renda_r, "vinc_r": vinc_r, "ql": ql})
        return rows

    def write_xlsx(cidade, level, rows, suffix):
        wb = Workbook()
        ws = wb.active
        ws.title = "Renda"
        nivel_lbl = {"secao": "SEÇÃO", "divisao": "DIVISÃO", "grupo": "GRUPO"}[level]
        ws.cell(1, 1, f"RENDA MÉDIA POR {nivel_lbl} CNAE (RAIS 2025) — {cidade.upper()} vs {REF.upper()}").font = title_font
        ws.cell(2, 1, "Vínculos ativos em 31/12, remuneração média nominal do ano (Vl Rem Média Nom). "
                      f"QL Renda = renda {cidade} / renda {REF}.")
        header_out = []
        if level in ("secao", "divisao", "grupo"):
            header_out += ["Seção", "Nome Seção"]
        if level in ("divisao", "grupo"):
            header_out += ["Divisão", "Nome Divisão"]
        if level == "grupo":
            header_out += ["Grupo", "Nome Grupo"]
        header_out += [f"Renda Média {cidade} (R$)", f"Vínculos {cidade}",
                        f"Renda Média {REF} (R$)", f"Vínculos {REF}", f"QL Renda ({cidade[:3]}/{REF[:2]})"]
        for j, h in enumerate(header_out):
            ws.cell(4, j + 1, h).font = bold

        r = 5
        for row in rows:
            idx = row["idx"]
            c = 1
            if level == "secao":
                ws.cell(r, c, idx); c += 1
                ws.cell(r, c, secoes_unicas.get(idx, "")); c += 1
            elif level == "divisao":
                info = divisoes_unicas[divisoes_unicas["Divisão"] == idx]
                ws.cell(r, c, info["Seção"].iloc[0] if len(info) else ""); c += 1
                ws.cell(r, c, info["Nome Seção"].iloc[0] if len(info) else ""); c += 1
                ws.cell(r, c, idx); c += 1
                ws.cell(r, c, info["Nome Divisão"].iloc[0] if len(info) else ""); c += 1
            else:
                info = grupo_lookup.get(idx, {})
                ws.cell(r, c, info.get("Seção", "")); c += 1
                ws.cell(r, c, info.get("Nome Seção", "")); c += 1
                ws.cell(r, c, info.get("Divisão", "")); c += 1
                ws.cell(r, c, info.get("Nome Divisão", "")); c += 1
                ws.cell(r, c, idx); c += 1
                ws.cell(r, c, info.get("Nome Grupo", "")); c += 1
            ws.cell(r, c, round(row["renda_c"], 6) if row["renda_c"] is not None else "—"); c += 1
            ws.cell(r, c, row["vinc_c"]); c += 1
            ws.cell(r, c, round(row["renda_r"], 6) if row["renda_r"] is not None else "—"); c += 1
            ws.cell(r, c, row["vinc_r"]); c += 1
            ws.cell(r, c, round(row["ql"], 6) if row["ql"] is not None else "—"); c += 1
            r += 1

        for j in range(1, len(header_out) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 16

        path = f"{OUT_DIR}/Renda_por_{suffix}_CNAE_RAIS_{cidade}.xlsx"
        wb.save(path)
        print("Gravado:", path)

    for cidade in MUNIC_CODES:
        for level, suffix in [("secao", "Secao"), ("divisao", "Divisao"), ("grupo", "Grupo")]:
            rows = build_level(cidade, level)
            write_xlsx(cidade, level, rows, suffix)

    print(f"CONCLUIDO — {time.time()-t0:.0f}s total")


if __name__ == "__main__":
    main()
