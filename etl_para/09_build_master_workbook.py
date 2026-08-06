"""Empresas_e_QL_{Cidade}_PA_BR.xlsx — workbook master que reúne, num só
arquivo, as abas finais de Empresas_Ativas_{Cidade}_completo.xlsx (6 abas)
e QL_por_Grupo_CNAE_{Cidade}.xlsx (4 abas), igual ao padrão de
Empresas_e_QL_Imperatriz_MA_BR.xlsx."""

from copy import copy

import openpyxl
from openpyxl import Workbook

OUT_DIR = "/Volumes/WISKET/painel-imperatriz-ma-main"
CIDADES = ["Ananindeua", "Capanema"]


def copy_sheet(src_ws, dst_wb, title):
    dst_ws = dst_wb.create_sheet(title=title[:31])
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
    return dst_ws


for cidade in CIDADES:
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    src1 = openpyxl.load_workbook(f"{OUT_DIR}/Empresas_Ativas_{cidade}_completo.xlsx", data_only=True)
    for name in src1.sheetnames:
        copy_sheet(src1[name], wb_out, name)

    src2 = openpyxl.load_workbook(f"{OUT_DIR}/QL_por_Grupo_CNAE_{cidade}.xlsx", data_only=True)
    for name in src2.sheetnames:
        copy_sheet(src2[name], wb_out, name)

    path = f"{OUT_DIR}/Empresas_e_QL_{cidade}_PA_BR.xlsx"
    wb_out.save(path)
    print("Gravado:", path, "—", len(wb_out.sheetnames), "abas:", wb_out.sheetnames)
