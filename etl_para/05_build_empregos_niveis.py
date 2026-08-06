"""Reconstrói Empregos_por_Secao/Divisao/Grupo_CNAE_RAIS_{Cidade}.xlsx a partir
do parquet já processado do RAIS Vínculos (2001-2025, 13 pontos), usando o
mapa oficial de Seção/Divisão/Grupo. QL sempre vs Pará (não há Brasil nesta
série). Métricas: QL por ano, QL Médio, Representat.(%), Cresc. 5 anos (%)
[valor[-4] -> valor[-1]], Consistência (0-12) [transições positivas]."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

STAGING = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/staging"
REFS = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/refs"
OUT_DIR = "/Volumes/WISKET/painel-imperatriz-ma-main"

ANOS = [2001, 2003, 2005, 2007, 2009, 2011, 2013, 2015, 2017, 2019, 2021, 2023, 2025]
MUNIC_CODES = {"Ananindeua": "150080", "Capanema": "150293"}
REF_CODE = "PARA"
REF = "Pará"

mapa = pd.read_csv(f"{REFS}/_mapa_grupo_secao_oficial.csv")
mapa["Grupo"] = mapa["Grupo"].astype(int)
mapa["Divisão"] = mapa["Divisão"].astype(int)
grupo_lookup = mapa.set_index("Grupo")[["Seção", "Nome Seção", "Divisão", "Nome Divisão", "Nome Grupo"]].to_dict("index")

secoes_unicas = mapa[["Seção", "Nome Seção"]].drop_duplicates().set_index("Seção")["Nome Seção"].to_dict()
divisoes_unicas = mapa[["Seção", "Nome Seção", "Divisão", "Nome Divisão"]].drop_duplicates()

df = pd.read_parquet(f"{STAGING}/rais_vinculos_agregado.parquet")
df = df.dropna(subset=["grupo_codigo"])
df["grupo_int"] = df["grupo_codigo"].astype(int)

bold = Font(bold=True)
title_font = Font(bold=True, size=12)


def pivot_counts(sub_df, group_keys):
    p = sub_df.pivot_table(index=group_keys, columns="ano", values="empregos", aggfunc="sum").fillna(0)
    return p.reindex(columns=ANOS, fill_value=0)


def calc_metrics(vals, total_last_year):
    """vals: lista de 13 valores (int). Retorna dict com QLs, medias, etc.
    QL é calculado fora (precisa do total da referência); aqui só as métricas
    que dependem só da série da própria cidade."""
    representat = (vals[-1] / total_last_year * 100) if total_last_year else None
    if vals[-4] not in (0, None):
        cresc5 = (vals[-1] - vals[-4]) / vals[-4] * 100
    else:
        cresc5 = None
    transicoes = sum(1 for i in range(1, len(vals)) if vals[i] > vals[i - 1])
    return representat, cresc5, f"{transicoes}/{len(vals)-1}"


def build_level(cidade, level):
    """level: 'secao' | 'divisao' | 'grupo'"""
    cidade_df = df[df.municipio == MUNIC_CODES[cidade]].copy()
    # Pará = resto (bucket 'PARA') + Ananindeua + Capanema (estado inteiro)
    para_total_df = df.groupby(["ano", "grupo_int"], as_index=False)["empregos"].sum()

    if level == "grupo":
        cidade_df["Seção"] = cidade_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Seção"))
        keys = ["grupo_int"]
    elif level == "divisao":
        cidade_df["divisao_int"] = cidade_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Divisão"))
        para_total_df["divisao_int"] = para_total_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Divisão"))
        keys = ["divisao_int"]
    else:
        cidade_df["Seção"] = cidade_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Seção"))
        para_total_df["Seção"] = para_total_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Seção"))
        keys = ["Seção"]

    cidade_piv = pivot_counts(cidade_df, keys)
    para_piv = pivot_counts(para_total_df, keys)

    cidade_total_by_year = cidade_piv.sum(axis=0)

    rows = []
    for idx in cidade_piv.index:
        vals = [int(cidade_piv.loc[idx, a]) for a in ANOS]
        ref_vals = [int(para_piv.loc[idx, a]) if idx in para_piv.index else 0 for a in ANOS]
        qls = []
        for v, rv, a in zip(vals, ref_vals, ANOS):
            city_tot = cidade_total_by_year[a]
            ref_tot = para_piv[a].sum()
            if city_tot == 0 or rv == 0 or ref_tot == 0:
                qls.append(None)
            else:
                qls.append((v / city_tot) / (rv / ref_tot))
        ql_validos = [q for q in qls if q is not None]
        ql_medio = sum(ql_validos) / len(ql_validos) if ql_validos else None
        representat, cresc5, consist = calc_metrics(vals, cidade_total_by_year[ANOS[-1]])
        rows.append({"idx": idx, "vals": vals, "qls": qls, "ql_medio": ql_medio,
                      "representat": representat, "cresc5": cresc5, "consist": consist})
    return rows


def write_xlsx(cidade, level, rows, filename_suffix):
    wb = Workbook()
    ws = wb.active
    ws.title = "Empregos"

    nivel_lbl = {"secao": "SEÇÃO", "divisao": "DIVISÃO", "grupo": "GRUPO"}[level]
    ws.cell(1, 1, f"EMPREGOS POR {nivel_lbl} CNAE (RAIS) — {cidade.upper()} vs {REF.upper()} — SÉRIE 2001-2025").font = title_font
    ws.cell(2, 1, "Vínculos empregatícios ativos em 31/12. Fonte: RAIS/MTE.")

    col = 1
    header = []
    if level in ("secao", "divisao", "grupo"):
        header += ["Seção", "Nome Seção"]
    if level in ("divisao", "grupo"):
        header += ["Divisão", "Nome Divisão"]
    if level == "grupo":
        header += ["Grupo", "Nome Grupo"]
    header += [f"Emp. {a}" for a in ANOS] + [f"QL {a}" for a in ANOS] + [
        "QL Médio", "Representat. (%)", "Cresc. 5 anos (%)", "Consistência (0-12)"
    ]
    for j, h in enumerate(header):
        ws.cell(3, j + 1, h).font = bold

    r = 4
    for row in sorted(rows, key=lambda x: (str(x["idx"]))):
        idx = row["idx"]
        c = 1
        if level == "secao":
            ws.cell(r, c, idx); c += 1
            ws.cell(r, c, secoes_unicas.get(idx, "")); c += 1
        elif level == "divisao":
            info = divisoes_unicas[divisoes_unicas["Divisão"] == idx]
            secao_c = info["Seção"].iloc[0] if len(info) else ""
            secao_n = info["Nome Seção"].iloc[0] if len(info) else ""
            divisao_n = info["Nome Divisão"].iloc[0] if len(info) else ""
            ws.cell(r, c, secao_c); c += 1
            ws.cell(r, c, secao_n); c += 1
            ws.cell(r, c, idx); c += 1
            ws.cell(r, c, divisao_n); c += 1
        else:  # grupo
            info = grupo_lookup.get(idx, {})
            ws.cell(r, c, info.get("Seção", "")); c += 1
            ws.cell(r, c, info.get("Nome Seção", "")); c += 1
            ws.cell(r, c, info.get("Divisão", "")); c += 1
            ws.cell(r, c, info.get("Nome Divisão", "")); c += 1
            ws.cell(r, c, idx); c += 1
            ws.cell(r, c, info.get("Nome Grupo", "")); c += 1

        for v in row["vals"]:
            ws.cell(r, c, v); c += 1
        for q in row["qls"]:
            ws.cell(r, c, round(q, 6) if q is not None else "—"); c += 1
        ws.cell(r, c, round(row["ql_medio"], 6) if row["ql_medio"] is not None else "—"); c += 1
        ws.cell(r, c, round(row["representat"], 6) if row["representat"] is not None else "—"); c += 1
        ws.cell(r, c, round(row["cresc5"], 6) if row["cresc5"] is not None else "—"); c += 1
        ws.cell(r, c, row["consist"]); c += 1
        r += 1

    for j in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14

    path = f"{OUT_DIR}/Empregos_por_{filename_suffix}_CNAE_RAIS_{cidade}.xlsx"
    wb.save(path)
    print("Gravado:", path)


for cidade in MUNIC_CODES:
    for level, suffix in [("secao", "Secao"), ("divisao", "Divisao"), ("grupo", "Grupo")]:
        rows = build_level(cidade, level)
        write_xlsx(cidade, level, rows, suffix)

print("CONCLUIDO")
