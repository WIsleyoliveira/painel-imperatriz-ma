"""Empresas_por_Secao/Divisao/Grupo_CNAE_RAIS_{Cidade}.xlsx — estoque de
empresas MATRIZ ativas hoje, reconstruído por data de abertura (corte 31/12
de cada ano), série 2001-2025 (13 pontos, mesma grade do RAIS). QL vs Pará.
Mesmas métricas que o de empregos: Representat.(%), Cresc. 5 anos (%),
Consistência (0-12)."""

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

STAGING = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/staging"
REFS = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/refs"
OUT_DIR = "/Volumes/WISKET/painel-imperatriz-ma-main"

ANOS = [2001, 2003, 2005, 2007, 2009, 2011, 2013, 2015, 2017, 2019, 2021, 2023, 2025]
MUNIC_CODES = {"Ananindeua": "0415", "Capanema": "0443"}
REF = "Pará"

mapa = pd.read_csv(f"{REFS}/_mapa_grupo_secao_oficial.csv")
mapa["Grupo"] = mapa["Grupo"].astype(int)
mapa["Divisão"] = mapa["Divisão"].astype(int)
grupo_lookup = mapa.set_index("Grupo")[["Seção", "Nome Seção", "Divisão", "Nome Divisão", "Nome Grupo"]].to_dict("index")
secoes_unicas = mapa[["Seção", "Nome Seção"]].drop_duplicates().set_index("Seção")["Nome Seção"].to_dict()
divisoes_unicas = mapa[["Seção", "Nome Seção", "Divisão", "Nome Divisão"]].drop_duplicates()

con = duckdb.connect()
con.execute("PRAGMA threads=6")

year_cols_ativas = ",\n".join(
    f"SUM(CASE WHEN ano_abertura<={y} AND is_ativa THEN 1 ELSE 0 END) AS ativas_{y}" for y in ANOS
)

print("Agregando por grupo (matriz, ativas, 13 anos)...", flush=True)
by_grupo = con.execute(f"""
    SELECT
        CASE WHEN municipio='0415' AND uf='PA' THEN 'Ananindeua'
             WHEN municipio='0443' AND uf='PA' THEN 'Capanema'
             WHEN uf='PA' THEN 'ParaResto'
             ELSE 'Outro' END AS geo_bucket,
        grupo_codigo,
        {year_cols_ativas}
    FROM read_parquet('{STAGING}/estabele_staging_v2_matriz.parquet')
    WHERE grupo_codigo IS NOT NULL AND uf='PA'
    GROUP BY 1, 2
""").df()
by_grupo["grupo_int"] = by_grupo["grupo_codigo"].astype(int)

bold = Font(bold=True)
title_font = Font(bold=True, size=12)


def geo_sum(df_, buckets, group_cols):
    sub = df_[df_["geo_bucket"].isin(buckets)]
    val_cols = [c for c in df_.columns if c.startswith("ativas_")]
    return sub.groupby(group_cols, as_index=False)[val_cols].sum()


PARA_BUCKETS = ["Ananindeua", "Capanema", "ParaResto"]


def calc_metrics(vals, total_last_year):
    representat = (vals[-1] / total_last_year * 100) if total_last_year else None
    cresc5 = (vals[-1] - vals[-4]) / vals[-4] * 100 if vals[-4] else None
    transicoes = sum(1 for i in range(1, len(vals)) if vals[i] > vals[i - 1])
    return representat, cresc5, f"{transicoes}/{len(vals)-1}"


def build_level(cidade, level):
    cidade_df = by_grupo[by_grupo.geo_bucket == cidade].copy()
    para_df = geo_sum(by_grupo, PARA_BUCKETS, ["grupo_int"])

    if level == "grupo":
        keys = ["grupo_int"]
    elif level == "divisao":
        cidade_df["divisao_int"] = cidade_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Divisão"))
        para_df["divisao_int"] = para_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Divisão"))
        keys = ["divisao_int"]
    else:
        cidade_df["Seção"] = cidade_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Seção"))
        para_df["Seção"] = para_df["grupo_int"].map(lambda g: grupo_lookup.get(g, {}).get("Seção"))
        keys = ["Seção"]

    val_cols = [f"ativas_{y}" for y in ANOS]
    cidade_piv = cidade_df.groupby(keys)[val_cols].sum().reindex(columns=val_cols, fill_value=0)
    para_piv = para_df.groupby(keys)[val_cols].sum().reindex(columns=val_cols, fill_value=0)

    cidade_total_by_year = cidade_piv.sum(axis=0)
    para_total_by_year = para_piv.sum(axis=0)

    rows = []
    for idx in cidade_piv.index:
        vals = [int(cidade_piv.loc[idx, f"ativas_{a}"]) for a in ANOS]
        ref_vals = [int(para_piv.loc[idx, f"ativas_{a}"]) if idx in para_piv.index else 0 for a in ANOS]
        qls = []
        for v, rv, a in zip(vals, ref_vals, ANOS):
            city_tot = cidade_total_by_year[f"ativas_{a}"]
            ref_tot = para_total_by_year[f"ativas_{a}"]
            if city_tot == 0 or rv == 0 or ref_tot == 0:
                qls.append(None)
            else:
                qls.append((v / city_tot) / (rv / ref_tot))
        ql_validos = [q for q in qls if q is not None]
        ql_medio = sum(ql_validos) / len(ql_validos) if ql_validos else None
        representat, cresc5, consist = calc_metrics(vals, cidade_total_by_year[f"ativas_{ANOS[-1]}"])
        rows.append({"idx": idx, "vals": vals, "qls": qls, "ql_medio": ql_medio,
                      "representat": representat, "cresc5": cresc5, "consist": consist})
    return rows


def write_xlsx(cidade, level, rows, filename_suffix):
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"

    nivel_lbl = {"secao": "SEÇÃO", "divisao": "DIVISÃO", "grupo": "GRUPO"}[level]
    nivel_cap = {"secao": "Seção", "divisao": "Divisão", "grupo": "Grupo"}[level]
    ws.cell(1, 1, f"EMPRESAS ATIVAS POR {nivel_lbl} CNAE (RECEITA FEDERAL) — {cidade.upper()} vs {REF.upper()} — SÉRIE 2001-2025").font = title_font
    ws.cell(2, 1, "Estoque de empresas (matriz) ativas hoje, reconstruído por data de abertura "
                  f"(corte 31/12 de cada ano). QL calculado no nível de {nivel_cap}. Fonte: CNPJ/Receita Federal.")

    header = []
    if level in ("secao", "divisao", "grupo"):
        header += ["Seção", "Nome Seção"]
    if level in ("divisao", "grupo"):
        header += ["Divisão", "Nome Divisão"]
    if level == "grupo":
        header += ["Grupo", "Nome Grupo"]
    header += [f"Emp. {a}" for a in ANOS] + [f"QL {a}" for a in ANOS] + [
        "QL Médio", "Representat. (%)", "Cresc. 5 anos (%)", "Consistência (0-12)"
    ]
    for j, h in enumerate(header):
        ws.cell(4, j + 1, h).font = bold

    r = 5
    for row in sorted(rows, key=lambda x: str(x["idx"])):
        idx = row["idx"]
        c = 1
        if level == "secao":
            ws.cell(r, c, idx); c += 1
            ws.cell(r, c, secoes_unicas.get(idx, "")); c += 1
        elif level == "divisao":
            info = divisoes_unicas[divisoes_unicas["Divisão"] == idx]
            ws.cell(r, c, info["Seção"].iloc[0] if len(info) else ""); c += 1
            ws.cell(r, c, info["Nome Seção"].iloc[0] if len(info) else ""); c += 1
            ws.cell(r, c, idx); c += 1
            ws.cell(r, c, info["Nome Divisão"].iloc[0] if len(info) else ""); c += 1
        else:
            info = grupo_lookup.get(idx, {})
            ws.cell(r, c, info.get("Seção", "")); c += 1
            ws.cell(r, c, info.get("Nome Seção", "")); c += 1
            ws.cell(r, c, info.get("Divisão", "")); c += 1
            ws.cell(r, c, info.get("Nome Divisão", "")); c += 1
            ws.cell(r, c, idx); c += 1
            ws.cell(r, c, info.get("Nome Grupo", "")); c += 1

        for v in row["vals"]:
            ws.cell(r, c, v); c += 1
        for q in row["qls"]:
            ws.cell(r, c, round(q, 6) if q is not None else "—"); c += 1
        ws.cell(r, c, round(row["ql_medio"], 6) if row["ql_medio"] is not None else "—"); c += 1
        ws.cell(r, c, round(row["representat"], 6) if row["representat"] is not None else "—"); c += 1
        ws.cell(r, c, round(row["cresc5"], 6) if row["cresc5"] is not None else "—"); c += 1
        ws.cell(r, c, row["consist"]); c += 1
        r += 1

    for j in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14

    path = f"{OUT_DIR}/Empresas_por_{filename_suffix}_CNAE_RAIS_{cidade}.xlsx"
    wb.save(path)
    print("Gravado:", path)


for cidade in MUNIC_CODES:
    for level, suffix in [("secao", "Secao"), ("divisao", "Divisao"), ("grupo", "Grupo")]:
        rows = build_level(cidade, level)
        write_xlsx(cidade, level, rows, suffix)

by_grupo.to_parquet(f"{STAGING}/agg_by_grupo_matriz_13anos.parquet")
print("CONCLUIDO")
