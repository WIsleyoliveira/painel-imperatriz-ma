"""A partir do staging parquet (estabelecimentos CNPJ, Brasil inteiro),
monta os 3 arquivos finais para Ananindeua e Capanema:

1. Empresas_Ativas_{Cidade}_completo.xlsx — 6 abas ({CIDADE}/PARÁ/BRASIL x
   Ativas/Todas), por Seção CNAE x Porte (MEI/Demais/Total) x Ano.
2. Estoque_Empresas_Ananindeua_Capanema.xlsx — 2 abas, estabelecimentos
   ativos por Seção CNAE x Ano, com Δ% (equivalente ao RAIS ESTAB, mas a
   partir do CNPJ, conforme decidido).
3. QL_por_Grupo_CNAE_{Cidade}.xlsx — 4 abas (ATIV_vs_PA/BR, TODA_vs_PA/BR),
   QL por Grupo CNAE x Ano, calculado aqui (não veio pronto desta vez).
"""

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

STAGING = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/staging"
REFS = "/Volumes/WISKET/painel-imperatriz-ma-main/etl_para/refs"
OUT_DIR = "/Volumes/WISKET/painel-imperatriz-ma-main"

ANOS = [2000, 2005, 2010, 2015, 2020, 2025]

secoes = pd.read_csv(f"{REFS}/cnae_secoes.csv", dtype=str).set_index("codigo")["nome"].to_dict()
grupo_secao = pd.read_csv(f"{REFS}/grupo_secao.csv", dtype=str)
grupo_nomes = grupo_secao.set_index("grupo_codigo")["grupo_nome"].to_dict()

CIDADES = {"Ananindeua": ("0415", "PA"), "Capanema": ("0443", "PA")}

con = duckdb.connect()
con.execute("PRAGMA threads=6")
con.execute(f"""
    CREATE VIEW staging AS
    SELECT *,
        CASE WHEN municipio='0415' AND uf='PA' THEN 'Ananindeua'
             WHEN municipio='0443' AND uf='PA' THEN 'Capanema'
             WHEN uf='PA' THEN 'ParaResto'
             ELSE 'BrasilResto' END AS geo_bucket
    FROM read_parquet('{STAGING}/estabele_staging.parquet')
""")

# ---------------------------------------------------------------------
# Query única: contagens por geo_bucket x secao x porte x ano-cutoff x tipo
# ---------------------------------------------------------------------
year_cols_ativas = ",\n".join(
    f"SUM(CASE WHEN ano_abertura<={y} AND is_ativa THEN 1 ELSE 0 END) AS ativas_{y}" for y in ANOS
)
year_cols_todas = ",\n".join(
    f"SUM(CASE WHEN ano_abertura<={y} THEN 1 ELSE 0 END) AS todas_{y}" for y in ANOS
)

print("Agregando por seção (para Empresas_Ativas e Estoque_Empresas)...", flush=True)
by_secao = con.execute(f"""
    SELECT geo_bucket, secao_codigo,
        CASE WHEN is_mei THEN 'MEI' ELSE 'Demais' END AS porte,
        {year_cols_ativas},
        {year_cols_todas}
    FROM staging
    WHERE secao_codigo IS NOT NULL
    GROUP BY 1,2,3
""").df()

print("Agregando por grupo (para QL)...", flush=True)
by_grupo = con.execute(f"""
    SELECT geo_bucket, grupo_codigo,
        {year_cols_ativas},
        {year_cols_todas}
    FROM staging
    WHERE grupo_codigo IS NOT NULL
    GROUP BY 1,2
""").df()

by_secao.to_parquet(f"{STAGING}/agg_by_secao.parquet")
by_grupo.to_parquet(f"{STAGING}/agg_by_grupo.parquet")
print("Agregações salvas. Montando xlsx...", flush=True)

# ---------------------------------------------------------------------
# Helper: soma de duas ou mais linhas geo_bucket (p.ex. Pará = cidade+cidade+resto)
# ---------------------------------------------------------------------
def geo_sum(df, buckets, group_cols):
    sub = df[df["geo_bucket"].isin(buckets)]
    val_cols = [c for c in df.columns if c.startswith("ativas_") or c.startswith("todas_")]
    return sub.groupby(group_cols, as_index=False)[val_cols].sum()


PARA_BUCKETS = ["Ananindeua", "Capanema", "ParaResto"]
BRASIL_BUCKETS = ["Ananindeua", "Capanema", "ParaResto", "BrasilResto"]

bold = Font(bold=True)
title_font = Font(bold=True, size=12)


def write_empresas_ativas(cidade, cod_munic):
    geo_data = {
        cidade: by_secao[by_secao["geo_bucket"] == cidade],
        "Pará": geo_sum(by_secao, PARA_BUCKETS, ["secao_codigo", "porte"]),
        "Brasil": geo_sum(by_secao, BRASIL_BUCKETS, ["secao_codigo", "porte"]),
    }
    wb = Workbook()
    wb.remove(wb.active)
    for geo_nome, df_geo in geo_data.items():
        for tipo in ["Ativas", "Todas"]:
            sheet_name = f"{geo_nome.upper()} — {tipo}"[:31]
            ws = wb.create_sheet(sheet_name)
            prefix = "ativas" if tipo == "Ativas" else "todas"
            titulo_tipo = "SOMENTE ATIVAS" if tipo == "Ativas" else "TODAS AS CRIADAS"
            ws.cell(1, 1, f"EMPRESAS POR PORTE E SEÇÃO CNAE — {geo_nome.upper()} — {titulo_tipo} — 2000 A 2025").font = title_font
            ws.cell(2, 1, "Fonte: CNPJ Receita Federal (extração Jul/2026). Porte limitado a MEI vs Demais "
                          "(arquivo EMPRESAS com detalhamento ME/EPP/Grande-Médio não disponível nesta base).")

            header_row = 3
            ws.cell(header_row, 1, "Seção")
            ws.cell(header_row, 2, "Descrição")
            col = 3
            for y in ANOS:
                ws.cell(header_row, col, y).font = bold
                col += 3
            sub_row = header_row + 1
            col = 3
            for y in ANOS:
                ws.cell(sub_row, col, "MEI")
                ws.cell(sub_row, col + 1, "Demais")
                ws.cell(sub_row, col + 2, "Total")
                col += 3

            secao_list = sorted(secoes.keys())
            r = sub_row + 1
            totals = {y: {"MEI": 0, "Demais": 0} for y in ANOS}
            for secao in secao_list:
                ws.cell(r, 1, secao)
                ws.cell(r, 2, secoes[secao])
                col = 3
                for y in ANOS:
                    mei_val = df_geo[(df_geo.secao_codigo == secao) & (df_geo.porte == "MEI")][f"{prefix}_{y}"].sum()
                    demais_val = df_geo[(df_geo.secao_codigo == secao) & (df_geo.porte == "Demais")][f"{prefix}_{y}"].sum()
                    ws.cell(r, col, int(mei_val))
                    ws.cell(r, col + 1, int(demais_val))
                    ws.cell(r, col + 2, int(mei_val + demais_val))
                    totals[y]["MEI"] += mei_val
                    totals[y]["Demais"] += demais_val
                    col += 3
                r += 1

            ws.cell(r, 1, "TOTAL").font = bold
            col = 3
            for y in ANOS:
                ws.cell(r, col, int(totals[y]["MEI"])).font = bold
                ws.cell(r, col + 1, int(totals[y]["Demais"])).font = bold
                ws.cell(r, col + 2, int(totals[y]["MEI"] + totals[y]["Demais"])).font = bold
                col += 3

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 55
            for c in range(3, 3 + 3 * len(ANOS)):
                ws.column_dimensions[get_column_letter(c)].width = 11

    path = f"{OUT_DIR}/Empresas_Ativas_{cidade}_completo.xlsx"
    wb.save(path)
    print("Gravado:", path)


def write_estoque_empresas():
    wb = Workbook()
    wb.remove(wb.active)
    geo_data = {
        "Ananindeua": by_secao[by_secao["geo_bucket"] == "Ananindeua"],
        "Capanema": by_secao[by_secao["geo_bucket"] == "Capanema"],
        "Pará": geo_sum(by_secao, PARA_BUCKETS, ["secao_codigo", "porte"]),
    }
    for geo_nome, df_geo in geo_data.items():
        ws = wb.create_sheet(geo_nome)
        ws.cell(1, 1, f"ESTOQUE DE ESTABELECIMENTOS (CNPJ ATIVOS) — {geo_nome.upper()} — 2000 A 2025").font = title_font
        ws.cell(2, 1, "Estabelecimentos com CNPJ ativo (situação 02) em Jul/2026, por data de abertura. "
                      "Fonte: CNPJ Receita Federal. Equivalente adaptado ao RAIS ESTAB (não baixado nesta série).")
        header_row = 3
        ws.cell(header_row, 1, "Seção").font = bold
        ws.cell(header_row, 2, "Descrição").font = bold
        for j, y in enumerate(ANOS):
            ws.cell(header_row, 3 + j, f"Estab.\n{y}").font = bold
        ws.cell(header_row, 3 + len(ANOS), "Δ%\n2000→2025").font = bold

        r = header_row + 1
        totals = [0] * len(ANOS)
        by_secao_only = df_geo.groupby("secao_codigo", as_index=False)[[f"ativas_{y}" for y in ANOS]].sum()
        for secao in sorted(secoes.keys()):
            row = by_secao_only[by_secao_only.secao_codigo == secao]
            vals = [int(row[f"ativas_{y}"].sum()) if len(row) else 0 for y in ANOS]
            ws.cell(r, 1, secao)
            ws.cell(r, 2, secoes[secao])
            for j, v in enumerate(vals):
                ws.cell(r, 3 + j, v)
                totals[j] += v
            if vals[0]:
                delta = (vals[-1] - vals[0]) / vals[0] * 100
                ws.cell(r, 3 + len(ANOS), f"{delta:.1f}%")
            else:
                ws.cell(r, 3 + len(ANOS), "—")
            r += 1
        ws.cell(r, 1, "TOTAL").font = bold
        for j, t in enumerate(totals):
            ws.cell(r, 3 + j, t).font = bold
        delta_t = (totals[-1] - totals[0]) / totals[0] * 100 if totals[0] else None
        ws.cell(r, 3 + len(ANOS), f"{delta_t:.1f}%" if delta_t is not None else "—").font = bold

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 55
        for j in range(len(ANOS) + 1):
            ws.column_dimensions[get_column_letter(3 + j)].width = 12

    path = f"{OUT_DIR}/Estoque_Empresas_Ananindeua_Capanema.xlsx"
    wb.save(path)
    print("Gravado:", path)


def compute_ql(cidade_df, ref_df, prefix):
    """QL por grupo para cada ano: (grupo/total cidade) / (grupo/total ref)."""
    cidade_tot = {y: cidade_df[f"{prefix}_{y}"].sum() for y in ANOS}
    ref_tot = {y: ref_df[f"{prefix}_{y}"].sum() for y in ANOS}
    grupos = sorted(set(cidade_df.grupo_codigo) | set(ref_df.grupo_codigo))
    rows = []
    for g in grupos:
        c_row = cidade_df[cidade_df.grupo_codigo == g]
        r_row = ref_df[ref_df.grupo_codigo == g]
        qls = []
        for y in ANOS:
            c_val = c_row[f"{prefix}_{y}"].sum() if len(c_row) else 0
            r_val = r_row[f"{prefix}_{y}"].sum() if len(r_row) else 0
            if cidade_tot[y] == 0 or r_val == 0 or ref_tot[y] == 0:
                qls.append(None)
            else:
                ql = (c_val / cidade_tot[y]) / (r_val / ref_tot[y])
                qls.append(ql)
        validos = [q for q in qls if q is not None]
        ql_medio = sum(validos) / len(validos) if validos else None
        periodos = sum(1 for i in range(1, len(qls)) if qls[i] is not None and qls[i - 1] is not None and qls[i] > qls[i - 1])
        rows.append([g, grupo_nomes.get(g, g)] + qls + [ql_medio, periodos])
    cols = ["codigo", "nome"] + [f"ql_{y}" for y in ANOS] + ["ql_medio", "periodos_cresc"]
    return pd.DataFrame(rows, columns=cols)


def write_ql(cidade):
    cidade_grupo = by_grupo[by_grupo.geo_bucket == cidade]
    para_grupo = geo_sum(by_grupo, PARA_BUCKETS, ["grupo_codigo"])
    brasil_grupo = geo_sum(by_grupo, BRASIL_BUCKETS, ["grupo_codigo"])

    sheets = {
        "ATIV_vs_PA": compute_ql(cidade_grupo, para_grupo, "ativas"),
        "ATIV_vs_BR": compute_ql(cidade_grupo, brasil_grupo, "ativas"),
        "TODA_vs_PA": compute_ql(cidade_grupo, para_grupo, "todas"),
        "TODA_vs_BR": compute_ql(cidade_grupo, brasil_grupo, "todas"),
    }
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name)
        tipo_lbl = "ATIVAS" if sheet_name.startswith("ATIV") else "TODAS"
        ref_lbl = "PARÁ" if sheet_name.endswith("PA") else "BRASIL"
        ws.cell(1, 1, f"QL POR GRUPO CNAE — {cidade.upper()} vs {ref_lbl} — {tipo_lbl} — 2000 A 2025").font = title_font
        ws.cell(3, 1, "Código").font = bold
        ws.cell(3, 2, "Nome do Grupo").font = bold
        for j, y in enumerate(ANOS):
            ws.cell(3, 3 + j, f"QL {y}").font = bold
        ws.cell(3, 3 + len(ANOS), "QL Médio").font = bold
        ws.cell(3, 4 + len(ANOS), "Períodos Cresc.").font = bold

        df_sorted = df.sort_values("ql_medio", ascending=False, na_position="last")
        r = 4
        for _, row in df_sorted.iterrows():
            ws.cell(r, 1, row["codigo"])
            ws.cell(r, 2, row["nome"])
            for j, y in enumerate(ANOS):
                v = row[f"ql_{y}"]
                ws.cell(r, 3 + j, round(v, 2) if v is not None else "—")
            ws.cell(r, 3 + len(ANOS), round(row["ql_medio"], 2) if row["ql_medio"] is not None else "—")
            ws.cell(r, 4 + len(ANOS), row["periodos_cresc"])
            r += 1

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 55
        for j in range(len(ANOS) + 2):
            ws.column_dimensions[get_column_letter(3 + j)].width = 12

    path = f"{OUT_DIR}/QL_por_Grupo_CNAE_{cidade}.xlsx"
    wb.save(path)
    print("Gravado:", path)


for cidade, (cod, uf) in CIDADES.items():
    write_empresas_ativas(cidade, cod)
    write_ql(cidade)
write_estoque_empresas()

print("CONCLUIDO")
