"""Aplica a formatação visual usada nos arquivos de referência (Imperatriz)
em todos os xlsx gerados para Ananindeua/Capanema: título azul #1E2A5E em
negrito, nota cinza, cabeçalho branco em negrito com fundo azul, linhas de
dados zebradas (F4F4F4/FFFFFF), congelamento de painel e larguras de coluna
razoáveis. Não mexe em nenhum valor, só estilo."""

import glob

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AZUL = "1E2A5E"
CINZA_NOTA = "555555"
ZEBRA_A = "F4F4F4"
ZEBRA_B = "FFFFFF"

OUT_DIR = "/Volumes/WISKET/painel-imperatriz-ma-main"

title_font = Font(bold=True, color=AZUL, size=12)
note_font = Font(color=CINZA_NOTA, italic=False)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor=AZUL)
fill_a = PatternFill("solid", fgColor=ZEBRA_A)
fill_b = PatternFill("solid", fgColor=ZEBRA_B)
data_font = Font(color="000000")
bold_id_font = Font(bold=True, color="000000")


def is_header_cell(v):
    if v is None:
        return False
    s = str(v).strip()
    return s in (
        "Seção", "Código", "Nível", "Porte"
    )


def find_header_row(ws, max_scan=8):
    for r in range(1, max_scan + 1):
        v = ws.cell(r, 1).value
        if is_header_cell(v):
            return r
    return None


def style_sheet(ws, id_cols_hint=2):
    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 2 or max_col < 1:
        return

    header_row = find_header_row(ws)
    if header_row is None:
        return

    # título (linha 1) — mescla e estiliza
    title_cell = ws.cell(1, 1)
    if title_cell.value:
        title_cell.font = title_font
        ws.row_dimensions[1].height = 30
        try:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        except Exception:
            pass

    # nota (linhas entre 2 e header_row-1, se houver texto)
    for r in range(2, header_row):
        note_cell = ws.cell(r, 1)
        if note_cell.value:
            note_cell.font = note_font
            ws.row_dimensions[r].height = 18
            try:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
            except Exception:
                pass

    # cabeçalho(s) — pode ter 2 linhas (ex: ano + porte nos arquivos antigos)
    header_rows = [header_row]
    # segunda linha de cabeçalho: se a linha logo abaixo ainda não parece dado
    # (primeira célula vazia/curta e sem ser letra de seção isolada), trata como subheader
    next_r = header_row + 1
    first_data_row = next_r
    if next_r <= max_row:
        c1 = ws.cell(next_r, 1).value
        row_vals = [ws.cell(next_r, c).value for c in range(2, min(max_col, 8) + 1)]
        has_text_val = any(v is not None and not isinstance(v, (int, float)) for v in row_vals)
        looks_like_subheader = (c1 in (None, "") and has_text_val)
        if looks_like_subheader:
            header_rows.append(next_r)
            first_data_row = next_r + 1

    for hr in header_rows:
        for c in range(1, max_col + 1):
            cell = ws.cell(hr, c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[hr].height = 30

    # linhas de dados — zebra
    band = 0
    for r in range(first_data_row, max_row + 1):
        if ws.cell(r, 1).value is None and ws.cell(r, 2).value is None:
            continue
        fill = fill_a if band % 2 == 0 else fill_b
        band += 1
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            if c <= id_cols_hint:
                cell.font = bold_id_font if c == 1 else data_font
            else:
                cell.font = data_font

    # larguras de coluna
    for c in range(1, max_col + 1):
        header_val = str(ws.cell(header_rows[0], c).value or "")
        letter = get_column_letter(c)
        if c == 1:
            width = 8
        elif "nome" in header_val.lower() or "descri" in header_val.lower():
            width = 45
        else:
            width = 12
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = ws.cell(first_data_row, 1).coordinate


def style_workbook(path):
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        id_hint = 6 if any(str(ws.cell(r, 5).value) == "Grupo" for r in range(1, 7)) else \
                  4 if any(str(ws.cell(r, 3).value) == "Divisão" for r in range(1, 7)) else 2
        style_sheet(ws, id_cols_hint=id_hint)
    wb.save(path)
    print("Formatado:", path)


CIDADES_ALVO = ["Ananindeua", "Capanema"]

FILES = sorted(set(
    glob.glob(f"{OUT_DIR}/Empregos_por_*_CNAE_RAIS_*.xlsx")
    + glob.glob(f"{OUT_DIR}/Empresas_por_*_CNAE_RAIS_*.xlsx")
    + glob.glob(f"{OUT_DIR}/Renda_por_*_CNAE_RAIS_*.xlsx")
    + glob.glob(f"{OUT_DIR}/QL_Grupo_Empresas_por_Porte_*.xlsx")
    + [f"{OUT_DIR}/QL_por_Grupo_CNAE_{c}.xlsx" for c in CIDADES_ALVO]
    + [f"{OUT_DIR}/Empresas_Ativas_{c}_completo.xlsx" for c in CIDADES_ALVO]
    + glob.glob(f"{OUT_DIR}/Estoque_Empregos_RAIS_Ananindeua_Capanema.xlsx")
    + glob.glob(f"{OUT_DIR}/Estoque_Empresas_Ananindeua_Capanema.xlsx")
    + glob.glob(f"{OUT_DIR}/Empresas_e_QL_*_PA_BR.xlsx")
))
FILES = [f for f in FILES if "Jul26" not in f and "com_nomes" not in f]

for f in FILES:
    try:
        style_workbook(f)
    except Exception as e:
        print("ERRO em", f, ":", e)

print(f"CONCLUIDO — {len(FILES)} arquivos processados")
